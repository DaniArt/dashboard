import os
import re
import json
from datetime import datetime
from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

DATA_FILE = os.path.join(os.path.dirname(__file__), 'data.json')
FILES_FILE = os.path.join(os.path.dirname(__file__), 'files.json')


def load_data():
    """Load data from JSON file on disk."""
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def load_files():
    """Load uploaded files registry."""
    if os.path.exists(FILES_FILE):
        with open(FILES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def save_data(data):
    """Save data to JSON file on disk."""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def save_files(files):
    """Save files registry to disk."""
    with open(FILES_FILE, 'w', encoding='utf-8') as f:
        json.dump(files, f, ensure_ascii=False, indent=2)


# Load persisted data on startup
DATA = load_data()
UPLOADED_FILES = load_files()

# FIO mapping: login -> full name
NAME_MAP = {
    'stashkin': 'Ташкин Санжар Булатович',
    'keabdrakhmanova': 'Абдрахманова Карлыгаш Ериковна',
    'kkarshiganov': 'Каршиганов Курмангали Исенович',
    'dbibol': 'Бибол Дәурен Дәулетұлы',
    'skan': 'Кан Сергей',
    'ykurbakovskiy': 'Курбаковский Юрий Евгеньевич',
    'rabayev': 'Абаев Рамазан Бауыржанович',
    'fsariyeva': 'Сариева Фариза Фатиховна',
    'kkaiyrbekov': 'Қайырбеков Қайырхан Амангелдіұлы',
    'asemiklet': 'Семиклет Александр Геннадьевич',
    'zfazylova': 'Фазылова Жансая Серікжанқызы',
    'ykhitrina': 'Дымочко Екатерина Сергеевна',
    'nnurakhmetov': 'Нурахметов Нуржан Айдынович',
    'arukassymova': 'Касымова Аружан Сакеновна',
    'rsagandykov': 'Сағандықов Рамин Әміржанұлы',
    'tnugumanov': 'Нугуманов Тимур Тимурланович',
    'amukanov': 'Муканов Аблайхан Омербекович',
    'kamirkhanov': 'Әмірханов Қуат Алтайұлы',
    'abaibulanov': 'Байбуланов Азамат Абилкасымович',
    'aistibayev': 'Истибаев Аман Ерсаинович',
    'asaubakirova': 'Аубакирова Айгерим Саметказыевна',
    'bkuanyshbayev': 'Куанышбаев Бауржан Имангалиевич',
    'zsatimbayev': 'Сәтімбаев Зәкіржан Алпамысұлы',
    'adrakhimova': 'Рахимова Ассоль Дамировна',
    'samussalimov': 'Мусалимов Шадияр Алимжанович',
    'tqsabyrov': 'Сабыров Теңелбек Қанатұлы',
    'imutemuratov': 'Утемуратов Иршод Муратулы',
    'ysamokhin': 'Самохин Евгений Викторович',
    'vbiriukov': 'Бирюков Владислав Викторович',
    'lvzhdanov': 'Жданов Леонид Валерьевич',
    'ibeskrovnov': 'Бескровнов Юрий Юрьевич',
    'dsagatov': 'Сағатов Данияр Талғатұлы',
    'dtugunov': 'Тугунов Дмитрий',
    'dkuatov': 'Куатов Данияр',
    'aaorazaliyev': 'Оразалиев Аблай Алтайұлы',
    'ivzhadan': 'Жадан Ирина Викторовна',
    'rkirilyuk': 'Кирилюк Руслан Дмитриевич',
    'rkinespaev': 'Кинеспаев Райнур Манарбекович',
}


def get_display_name(login):
    """Get full name from login, case-insensitive."""
    key = login.lower().strip()
    # Remove email parts if present
    key = key.split('@')[0]
    # Handle multi-word usernames (take first word as login)
    # e.g. "IMUtemuratov imutemuratov@fortebank.com" or "arukassymova arukassymovaNewpass..."
    if ' ' in key:
        key = key.split()[0]
    # Remove password/garbage suffixes (e.g. "arukassymovanewpass201823624!")
    for known in NAME_MAP:
        if key.startswith(known):
            key = known
            break
    return NAME_MAP.get(key, login)


def parse_timestamps_from_cell(cell_value):
    """Extract all timestamps from a cell value like '2026.05.18 09:01 2026.05.18 09:02 ...'"""
    if not cell_value:
        return []
    text = str(cell_value).strip()
    # Pattern: YYYY.MM.DD HH:MM
    pattern = r'(\d{4}\.\d{2}\.\d{2}\s+\d{2}:\d{2})'
    matches = re.findall(pattern, text)
    timestamps = []
    for m in matches:
        try:
            dt = datetime.strptime(m, '%Y.%m.%d %H:%M')
            timestamps.append(dt)
        except ValueError:
            continue
    return timestamps


def parse_xlsx(filepath):
    """Parse VPN export xlsx and return list of records. Supports two formats."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    records = []

    # Detect format by checking headers
    first_rows = []
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if row:
            first_rows.append(row)

    # Check for new Splunk-style format: columns _time, User_Name, Acct_Session_Id, Acct_Status_Type
    is_splunk_format = False
    for row in first_rows:
        row_text = ' '.join(str(c or '') for c in row).lower()
        if '_time' in row_text and 'user_name' in row_text:
            is_splunk_format = True
            break

    if is_splunk_format:
        records = parse_splunk_format(ws)
    else:
        records = parse_fortigate_format(ws, first_rows)

    wb.close()
    return records


def parse_splunk_format(ws):
    """Parse Splunk-style VPN export: _time, User_Name, Acct_Session_Id, Acct_Status_Type, count.
    Groups by session ID to correctly handle overnight sessions."""
    records = []
    # Find column indices from header row
    header = None
    header_row_idx = 1
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), start=1):
        if row:
            cells = [str(c or '').strip() for c in row]
            if '_time' in cells or any('time' in c.lower() for c in cells):
                header = cells
                header_row_idx = idx
                break

    if not header:
        return records

    # Map columns
    time_col = next((i for i, c in enumerate(header) if 'time' in c.lower()), None)
    user_col = next((i for i, c in enumerate(header) if 'user' in c.lower()), None)
    session_col = next((i for i, c in enumerate(header) if 'session' in c.lower()), None)
    status_col = next((i for i, c in enumerate(header) if 'status' in c.lower()), None)

    if time_col is None or user_col is None:
        return records

    # Collect all events
    events = []  # list of (datetime, username, session_id, status)
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or len(row) <= max(time_col, user_col):
            continue

        time_val = row[time_col]
        user_val = row[user_col]
        session_id = str(row[session_col]).strip() if session_col is not None and len(row) > session_col and row[session_col] else None
        status = str(row[status_col]).strip().lower() if status_col is not None and len(row) > status_col and row[status_col] else ''

        if not time_val or not user_val:
            continue

        time_str = str(time_val).strip()
        username = str(user_val).strip()

        if not username or username.lower() in ('user_name', ''):
            continue

        # Parse ISO timestamp
        ts = None
        try:
            clean = re.sub(r'\.\d+[+-]\d{2}:\d{2}$', '', time_str)
            ts = datetime.strptime(clean, '%Y-%m-%dT%H:%M:%S')
        except ValueError:
            try:
                clean = re.sub(r'\.\d+.*$', '', time_str)
                ts = datetime.strptime(clean, '%Y-%m-%dT%H:%M:%S')
            except ValueError:
                continue

        if ts:
            events.append((ts, username, session_id, status))

    # If we have session IDs and status, group by session
    if session_col is not None and status_col is not None:
        # Group events by session_id
        sessions = {}  # session_id -> {user, starts: [], stops: []}
        for ts, username, session_id, status in events:
            if not session_id:
                continue
            if session_id not in sessions:
                sessions[session_id] = {'user': username, 'starts': [], 'stops': []}
            if 'start' in status:
                sessions[session_id]['starts'].append(ts)
            elif 'stop' in status:
                sessions[session_id]['stops'].append(ts)

        # For each session, record it on the START day with start time and stop time
        for sid, data in sessions.items():
            username = data['user']
            start_ts = min(data['starts']) if data['starts'] else None
            stop_ts = max(data['stops']) if data['stops'] else None

            if start_ts:
                # Session belongs to the day it STARTED
                session_date = start_ts.strftime('%Y-%m-%d')
                start_time = start_ts.strftime('%H:%M')
                end_time = stop_ts.strftime('%H:%M') if stop_ts else start_time

                # If stop is next day, cap at 23:59 for display but keep for hours calc
                # Actually, just record the real end time — aggregate will take max
                records.append({
                    'username': username,
                    'display_name': get_display_name(username),
                    'date': session_date,
                    'start': start_time,
                    'end': end_time,
                })
    else:
        # No session tracking — fall back to simple per-event records
        for ts, username, session_id, status in events:
            records.append({
                'username': username,
                'display_name': get_display_name(username),
                'date': ts.strftime('%Y-%m-%d'),
                'start': ts.strftime('%H:%M'),
                'end': ts.strftime('%H:%M'),
            })

    return records


def parse_fortigate_format(ws, first_rows):
    """Parse original FortiGate VPN export with timestamps in one cell."""
    records = []

    # Validate: check that the file looks like a VPN export
    header_found = False
    for row in first_rows:
        row_text = ' '.join(str(c or '') for c in row).lower()
        if any(kw in row_text for kw in ['пользователь', 'сессия', 'nai-policy', 'статус аутентификации', 'номер сессии']):
            header_found = True
            break

    if not header_found:
        raise ValueError('Файл не похож на выгрузку VPN. Убедитесь что загружаете правильный файл.')

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue

        # Column 0: timestamps, Column 2: username
        timestamps_cell = row[0]
        username = row[2] if len(row) > 2 else None

        if not username or not timestamps_cell:
            continue

        # Skip mobile sessions (Android/iOS) — column 10 is OS
        if len(row) > 10 and row[10]:
            os_val = str(row[10]).strip().lower()
            if os_val in ('android', 'ios') or 'android' in os_val or 'ios' in os_val or 'iphone' in os_val:
                continue

        username = str(username).strip()
        if not username or username.lower() in ('пользователь', 'user', ''):
            continue

        timestamps = parse_timestamps_from_cell(timestamps_cell)
        if not timestamps:
            # Try alternate date format: M/D/YY H:MM
            text = str(timestamps_cell).strip()
            alt_pattern = r'(\d{1,2}/\d{1,2}/\d{2}\s+\d{1,2}:\d{2})'
            alt_matches = re.findall(alt_pattern, text)
            for m in alt_matches:
                try:
                    dt = datetime.strptime(m, '%m/%d/%y %H:%M')
                    timestamps.append(dt)
                except ValueError:
                    continue

        if not timestamps:
            continue

        first_ts = min(timestamps)
        last_ts = max(timestamps)

        records.append({
            'username': username,
            'display_name': get_display_name(username),
            'date': first_ts.strftime('%Y-%m-%d'),
            'start': first_ts.strftime('%H:%M'),
            'end': last_ts.strftime('%H:%M'),
        })

    return records


def aggregate_by_user_day(records):
    """Aggregate: per user per day, take earliest start and latest end."""
    agg = {}
    for r in records:
        key = (r['display_name'], r['date'])
        if key not in agg:
            agg[key] = {'start': r['start'], 'end': r['end']}
        else:
            if r['start'] < agg[key]['start']:
                agg[key]['start'] = r['start']
            if r['end'] > agg[key]['end']:
                agg[key]['end'] = r['end']

    result = []
    for (name, date), times in sorted(agg.items()):
        result.append({
            'name': name,
            'date': date,
            'start': times['start'],
            'end': times['end'],
        })
    return result


@app.route('/')
def index():
    return render_template('index.html')


# Kazakhstan 2026 public holidays (non-working days)
KZ_HOLIDAYS_2026 = {
    '2026-01-01', '2026-01-02', '2026-01-07',  # New Year + Orthodox Christmas
    '2026-03-08', '2026-03-09',  # Women's Day + moved
    '2026-03-21', '2026-03-22', '2026-03-23', '2026-03-24', '2026-03-25',  # Nauryz
    '2026-05-01',  # Unity Day
    '2026-05-07',  # Defender's Day
    '2026-05-09', '2026-05-11',  # Victory Day + moved
    '2026-05-27',  # Kurban Ait
    '2026-07-06',  # Capital City Day
    '2026-08-30', '2026-08-31',  # Constitution Day + moved
    '2026-10-25', '2026-10-26',  # Republic Day + moved
    '2026-12-16', '2026-12-17',  # Independence Day
}


def is_workday(date_str):
    """Check if a date is a working day in Kazakhstan (Mon-Fri, not a holiday)."""
    from datetime import date as dt_date
    d = dt_date.fromisoformat(date_str)
    # Weekend
    if d.weekday() >= 5:  # Saturday=5, Sunday=6
        return False
    # Public holiday
    if date_str in KZ_HOLIDAYS_2026:
        return False
    return True


def is_weekend_or_holiday(date_str):
    """Check if date is weekend or public holiday."""
    from datetime import date as dt_date
    d = dt_date.fromisoformat(date_str)
    if d.weekday() >= 5:
        return 'weekend'
    if date_str in KZ_HOLIDAYS_2026:
        return 'holiday'
    return None


@app.route('/api/calendar')
def get_calendar():
    """Return list of holidays for frontend."""
    return jsonify(list(KZ_HOLIDAYS_2026))


@app.route('/api/upload', methods=['POST'])
def upload():
    """Upload xlsx file, parse it, add to DATA."""
    if 'file' not in request.files:
        return jsonify({'error': 'Нет файла'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Пустое имя файла'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Нужен .xlsx файл'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)

    try:
        records = parse_xlsx(filepath)
        if not records:
            return jsonify({'error': 'Файл распознан, но записи VPN-сессий не найдены. Проверьте формат.'}), 400
        aggregated = aggregate_by_user_day(records)

        # Assign file_id to each record
        file_id = file.filename + '_' + datetime.now().strftime('%Y%m%d%H%M%S')
        for r in aggregated:
            r['file_id'] = file_id

        DATA.extend(aggregated)
        save_data(DATA)

        # Register file
        UPLOADED_FILES.append({'id': file_id, 'filename': file.filename, 'records': len(aggregated), 'uploaded_at': datetime.now().strftime('%Y-%m-%d %H:%M')})
        save_files(UPLOADED_FILES)

        return jsonify({
            'success': True,
            'message': f'Загружено {len(aggregated)} записей из {file.filename}',
            'total_records': len(DATA)
        })
    except Exception as e:
        return jsonify({'error': f'Ошибка парсинга: {str(e)}'}), 500


@app.route('/api/data')
def get_data():
    """Return all parsed data as JSON, with missing workdays filled in and weekend flags."""
    from datetime import date as dt_date, timedelta

    if not DATA:
        return jsonify([])

    # Group by user
    user_data = {}
    for r in DATA:
        name = r['name']
        if name not in user_data:
            user_data[name] = {}
        user_data[name][r['date']] = r

    # Find global date range from data
    all_dates = [r['date'] for r in DATA]
    min_date = dt_date.fromisoformat(min(all_dates))
    max_date = dt_date.fromisoformat(max(all_dates))

    # Build full result with missing days and flags
    result = []
    for name in sorted(user_data.keys()):
        current = min_date
        while current <= max_date:
            ds = current.isoformat()
            day_type = is_weekend_or_holiday(ds)

            if ds in user_data[name]:
                # Has data for this day
                entry = dict(user_data[name][ds])
                if day_type == 'weekend':
                    entry['flag'] = 'weekend_work'
                elif day_type == 'holiday':
                    entry['flag'] = 'holiday_work'
                else:
                    entry['flag'] = 'normal'
                result.append(entry)
            else:
                # No data
                if day_type is None:
                    # It's a workday but no VPN record = absent
                    result.append({
                        'name': name,
                        'date': ds,
                        'start': '',
                        'end': '',
                        'flag': 'absent'
                    })
                elif day_type == 'weekend':
                    # Weekend — show as rest day
                    result.append({
                        'name': name,
                        'date': ds,
                        'start': '',
                        'end': '',
                        'flag': 'weekend_off'
                    })
                elif day_type == 'holiday':
                    # Public holiday — show as rest day
                    result.append({
                        'name': name,
                        'date': ds,
                        'start': '',
                        'end': '',
                        'flag': 'holiday_off'
                    })

            current += timedelta(days=1)

    return jsonify(result)


@app.route('/api/files')
def get_files():
    """Return list of uploaded files."""
    return jsonify(UPLOADED_FILES)


@app.route('/api/files/<path:file_id>', methods=['DELETE'])
def delete_file(file_id):
    """Delete all records from a specific file."""
    before = len(DATA)
    DATA[:] = [r for r in DATA if r.get('file_id') != file_id]
    UPLOADED_FILES[:] = [f for f in UPLOADED_FILES if f['id'] != file_id]
    save_data(DATA)
    save_files(UPLOADED_FILES)
    removed = before - len(DATA)
    return jsonify({'success': True, 'message': f'Удалено {removed} записей', 'total_records': len(DATA)})
def clear_data():
    """Clear all data."""
    DATA.clear()
    save_data(DATA)
    return jsonify({'success': True, 'message': 'Данные очищены'})


if __name__ == '__main__':
    app.run(debug=False, host='127.0.0.1', port=5000)
